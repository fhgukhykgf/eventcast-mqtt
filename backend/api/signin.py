"""
签到管理接口
"""
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from datetime import datetime
import logging
import io
from typing import Dict, Any, Optional

from utils.database import get_database
from utils.mqtt_client import publish_message, is_mqtt_connected
from utils.auth import get_current_user, require_organizer, TokenData
from models.signin import ApplyRequest, SignInRequest

router = APIRouter()
logger = logging.getLogger(__name__)


@router.post("/apply")
async def apply_event(apply_data: ApplyRequest, current_user: TokenData = Depends(get_current_user)):
    """报名活动（需登录，从Token获取身份防止冒用）"""
    try:
        db = await get_database()
        events = db["events"]
        applies = db["user_apply"]

        # 安全：从 Token 获取用户身份，防止冒用
        real_user_id = current_user.user_id
        # 从数据库获取真实用户名
        user_doc = await db["users"].find_one({"user_id": real_user_id})
        real_user_name = user_doc.get("real_name", "") if user_doc else real_user_id

        event = await events.find_one({"event_id": apply_data.event_id})
        if not event:
            raise HTTPException(status_code=404, detail="活动不存在")

        if event["status"] != "active":
            raise HTTPException(status_code=400, detail="活动已结束")

        existing = await applies.find_one({
            "event_id": apply_data.event_id,
            "user_id": real_user_id
        })
        if existing:
            raise HTTPException(status_code=400, detail="已报名")

        apply_record = {
            "event_id": apply_data.event_id,
            "user_id": real_user_id,
            "user_name": real_user_name,
            "apply_time": datetime.now().isoformat(),
            "status": "applied"
        }

        try:
            await applies.insert_one(apply_record)
        except Exception as e:
            if "duplicate key error" in str(e).lower() or "e11000" in str(e).lower():
                raise HTTPException(status_code=400, detail="已报名")
            raise HTTPException(status_code=500, detail=str(e))

        # 安全地递增报名人数（避免并发超卖和负数）
        if event.get("limit_num"):
            result = await events.update_one(
                {
                    "event_id": apply_data.event_id,
                    "apply_count": {"$lt": event["limit_num"]}
                },
                {"$inc": {"apply_count": 1}}
            )
            if result.modified_count == 0:
                # 回滚报名记录
                await applies.delete_one({"_id": apply_record.get("_id")})
                raise HTTPException(status_code=400, detail="报名人数已满")
        else:
            await events.update_one(
                {"event_id": apply_data.event_id, "apply_count": {"$gte": 0}},
                {"$inc": {"apply_count": 1}}
            )

        try:
            if is_mqtt_connected():
                publish_message(f"event/{apply_data.event_id}/notice", {
                    "type": "apply_success",
                    "user_id": real_user_id,
                    "user_name": real_user_name,
                    "apply_time": apply_record["apply_time"]
                })
        except Exception as mqtt_error:
            logger.warning(f"MQTT通知发送失败: {mqtt_error}")

        logger.info(f"✅ 报名成功: {real_user_id} -> {apply_data.event_id}")
        return {"code": 200, "msg": "报名成功"}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ 报名失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/cancel")
async def cancel_apply(apply_data: ApplyRequest, current_user: TokenData = Depends(get_current_user)):
    """取消报名（需登录，从Token获取身份防止冒用）"""
    try:
        # 安全：从 Token 获取用户身份
        real_user_id = current_user.user_id

        db = await get_database()
        events = db["events"]
        applies = db["user_apply"]

        result = await applies.delete_one({
            "event_id": apply_data.event_id,
            "user_id": real_user_id
        })

        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="报名记录不存在")

        # 防止 apply_count 变为负数
        await events.update_one(
            {"event_id": apply_data.event_id, "apply_count": {"$gt": 0}},
            {"$inc": {"apply_count": -1}}
        )

        logger.info(f"✅ 取消报名: {real_user_id} -> {apply_data.event_id}")
        return {"code": 200, "msg": "取消成功"}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ 取消失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/in")
async def sign_in(sign_data: SignInRequest, current_user: TokenData = Depends(get_current_user)):
    """签到（需登录，从Token获取身份防止冒用）"""
    try:
        # 安全：从 Token 获取用户身份，防止冒用
        real_user_id = current_user.user_id
        # 从数据库获取真实用户名
        db = await get_database()
        user_doc = await db["users"].find_one({"user_id": real_user_id})
        real_user_name = user_doc.get("real_name", "") if user_doc else real_user_id

        events = db["events"]
        applies = db["user_apply"]
        signs = db["sign_records"]

        event = await events.find_one({"event_id": sign_data.event_id})
        if not event:
            raise HTTPException(status_code=404, detail="活动不存在")

        apply = await applies.find_one({
            "event_id": sign_data.event_id,
            "user_id": real_user_id
        })
        if not apply:
            raise HTTPException(status_code=400, detail="请先报名")

        existing = await signs.find_one({
            "event_id": sign_data.event_id,
            "user_id": real_user_id
        })
        if existing:
            raise HTTPException(status_code=400, detail="已签到")

        sign_record = {
            "event_id": sign_data.event_id,
            "user_id": real_user_id,
            "user_name": real_user_name,
            "sign_time": datetime.now().isoformat(),
            "sign_method": sign_data.sign_method or "scan"
        }

        try:
            await signs.insert_one(sign_record)
        except Exception as e:
            if "duplicate key error" in str(e).lower() or "e11000" in str(e).lower():
                raise HTTPException(status_code=400, detail="已签到")
            raise HTTPException(status_code=500, detail=str(e))

        await events.update_one(
            {"event_id": sign_data.event_id},
            {"$inc": {"sign_count": 1}}
        )

        sign_rate = 0
        if event["apply_count"] > 0:
            sign_rate = round((event["sign_count"] + 1) / event["apply_count"] * 100, 2)

        try:
            if is_mqtt_connected():
                publish_message(f"event/{sign_data.event_id}/sign_in", {
                    "type": "sign_in",
                    "user_id": real_user_id,
                    "user_name": real_user_name,
                    "sign_time": sign_record["sign_time"]
                })
        except Exception as mqtt_error:
            logger.warning(f"MQTT通知发送失败: {mqtt_error}")

        logger.info(f"✅ 签到成功: {real_user_id} -> {sign_data.event_id}")
        return {
            "code": 200,
            "msg": "签到成功",
            "data": {
                "sign_time": sign_record["sign_time"],
                "sign_rate": f"{sign_rate}%"
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ 签到失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/status/{event_id}/{user_id}")
async def get_sign_status(event_id: str, user_id: str, current_user: TokenData = Depends(get_current_user)):
    """获取签到状态（需登录，仅可查看自己或管理员/组织者可查看所有人）"""
    try:
        # 访问控制：只能查看自己的状态，除非是管理员或组织者
        if user_id != current_user.user_id and current_user.role not in ["admin", "organizer"]:
            raise HTTPException(status_code=403, detail="无权查看其他用户签到状态")

        db = await get_database()

        apply = await db["user_apply"].find_one({
            "event_id": event_id,
            "user_id": user_id
        })

        sign = await db["sign_records"].find_one({
            "event_id": event_id,
            "user_id": user_id
        })

        return {
            "code": 200,
            "data": {
                "has_applied": apply is not None,
                "has_signed": sign is not None,
                "apply_time": apply.get("apply_time") if apply else None,
                "sign_time": sign.get("sign_time") if sign else None
            }
        }

    except Exception as e:
        logger.error(f"获取签到状态失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/records/{event_id}")
async def get_sign_records(event_id: str, skip: int = 0, limit: int = 50, current_user: TokenData = Depends(require_organizer)):
    """获取活动签到记录（需组织者或管理员权限）"""
    try:
        db = await get_database()

        total = await db["sign_records"].count_documents({"event_id": event_id})

        cursor = db["sign_records"].find(
            {"event_id": event_id}
        ).sort("sign_time", -1).skip(skip).limit(limit)

        records = await cursor.to_list(length=limit)

        result = []
        for r in records:
            result.append({
                "user_id": r["user_id"],
                "user_name": r["user_name"],
                "sign_time": r["sign_time"],
                "sign_method": r.get("sign_method", "scan")
            })

        return {"code": 200, "data": result, "total": total}

    except Exception as e:
        logger.error(f"获取签到记录失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/count/{event_id}")
async def get_sign_count(event_id: str):
    """获取签到计数（公开接口）"""
    try:
        db = await get_database()

        event = await db["events"].find_one({"event_id": event_id})
        if not event:
            raise HTTPException(status_code=404, detail="活动不存在")

        sign_rate = 0
        if event["apply_count"] > 0:
            sign_rate = round(event["sign_count"] / event["apply_count"] * 100, 2)

        return {
            "code": 200,
            "data": {
                "apply_count": event["apply_count"],
                "sign_count": event["sign_count"],
                "sign_rate": f"{sign_rate}%"
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取签到计数失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/user/{user_id}")
async def get_user_sign_records(user_id: str, skip: int = 0, limit: int = 20, current_user: TokenData = Depends(get_current_user)):
    """
    获取用户的活动记录（需登录）
    """
    try:
        db = await get_database()

        applies = await db["user_apply"].find(
            {"user_id": user_id}
        ).sort("apply_time", -1).skip(skip).limit(limit).to_list(length=limit)

        signs = await db["sign_records"].find(
            {"user_id": user_id}
        ).to_list(length=None)

        sign_map = {s["event_id"]: s for s in signs}

        result = []
        for apply in applies:
            event = await db["events"].find_one({"event_id": apply["event_id"]})
            if event:
                sign_record = sign_map.get(apply["event_id"])
                
                # 计算活动状态
                event_status = event.get("status", "active")
                if event_status != "cancelled":
                    now = datetime.now()
                    end_time_str = event.get("end_time") or event.get("time")
                    if end_time_str:
                        try:
                            end_time = datetime.fromisoformat(end_time_str.replace('Z', '+00:00').replace(' ', 'T'))
                            if now > end_time:
                                event_status = "ended"
                        except (ValueError, TypeError) as e:
                            logger.debug(f"解析活动结束时间失败: {end_time_str}, 错误: {e}")
                
                result.append({
                    "event_id": apply["event_id"],
                    "event_title": event.get("title", ""),
                    "event_time": event.get("time", ""),
                    "event_start_time": event.get("start_time", ""),
                    "event_end_time": event.get("end_time", ""),
                    "event_location": event.get("location", ""),
                    "event_status": event_status,
                    "apply_time": apply.get("apply_time", ""),
                    "sign_time": sign_record.get("sign_time") if sign_record else None,
                    "status": "signed" if sign_record else "applied"
                })

        total = await db["user_apply"].count_documents({"user_id": user_id})

        return {
            "code": 200,
            "data": result,
            "total": total
        }

    except Exception as e:
        logger.error(f"获取用户活动记录失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/all/{event_id}")
async def get_all_applicants(event_id: str, skip: int = 0, limit: int = 50, search: Optional[str] = None, status: Optional[str] = None, current_user: TokenData = Depends(require_organizer)):
    """
    获取活动的所有报名人员（需组织者或管理员权限）
    status 参数: 'applied' (未签到), 'signed' (已签到)
    """
    try:
        db = await get_database()

        query: Dict[str, Any] = {"event_id": event_id}
        if search:
            # 安全：转义正则特殊字符防止 ReDoS，限制搜索长度
            import re as _re
            safe_search = _re.escape(search[:50])
            query["$or"] = [
                {"user_name": {"$regex": safe_search, "$options": "i"}},
                {"user_id": {"$regex": safe_search, "$options": "i"}}
            ]

        total = await db["user_apply"].count_documents(query)

        cursor = db["user_apply"].find(query).sort("apply_time", -1).skip(skip).limit(limit)
        applies = await cursor.to_list(length=limit)

        signs = await db["sign_records"].find(
            {"event_id": event_id}
        ).to_list(length=None)

        sign_map = {s["user_id"]: s for s in signs}

        result = []
        for apply in applies:
            sign_record = sign_map.get(apply["user_id"])
            record_status = "signed" if sign_record else "applied"
            
            # 如果指定了 status 过滤，则只返回匹配的记录
            if status and record_status != status:
                continue
            
            result.append({
                "user_id": apply["user_id"],
                "user_name": apply["user_name"],
                "apply_time": apply["apply_time"],
                "sign_time": sign_record.get("sign_time") if sign_record else None,
                "status": record_status
            })

        # 如果有 status 过滤，需要重新计算 total
        if status:
            total = len(result)

        return {
            "code": 200,
            "data": result,
            "total": total
        }

    except Exception as e:
        logger.error(f"获取报名人员失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export/{event_id}")
async def export_sign_list(event_id: str, current_user: TokenData = Depends(require_organizer)):
    """
    导出活动签到名单为 Excel（需组织者或管理员权限）
    """
    try:
        # 检查 openpyxl 是否可用
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            raise HTTPException(status_code=500, detail="服务器未安装 openpyxl 库，请联系管理员")
        
        db = await get_database()
        
        # 获取活动信息
        event = await db["events"].find_one({"event_id": event_id})
        if not event:
            raise HTTPException(status_code=404, detail="活动不存在")
        
        # 获取所有报名人员
        applies_cursor = db["user_apply"].find({"event_id": event_id}).sort("apply_time", 1)
        applies = await applies_cursor.to_list(length=None)
        
        # 获取所有签到记录
        signs = await db["sign_records"].find({"event_id": event_id}).to_list(length=None)
        sign_map = {s["user_id"]: s for s in signs}
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "签到名单"
        
        # 设置样式
        title_font = Font(name='微软雅黑', size=16, bold=True, color='1890FF')
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1890FF', end_color='1890FF', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 标题行
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = f"{event.get('title', '')} - 签到名单"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 40
        
        # 活动信息
        ws.merge_cells('A2:E2')
        info_cell = ws['A2']
        start_time = event.get('start_time') or event.get('time', '')
        end_time = event.get('end_time', '')
        info_cell.value = f"时间：{start_time} 至 {end_time}  地点：{event.get('location', '')}  组织者：{event.get('organizer', '')}"
        info_cell.font = Font(name='微软雅黑', size=10, color='666666')
        info_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[2].height = 25
        
        # 统计信息
        ws.merge_cells('A3:E3')
        stats_cell = ws['A3']
        apply_count = event.get('apply_count', 0)
        sign_count = event.get('sign_count', 0)
        sign_rate = f"{round(sign_count / apply_count * 100, 2)}%" if apply_count > 0 else "0%"
        stats_cell.value = f"报名人数：{apply_count} 人  签到人数：{sign_count} 人  签到率：{sign_rate}"
        stats_cell.font = Font(name='微软雅黑', size=10, bold=True, color='52c41a')
        stats_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[3].height = 25
        
        # 表头
        headers = ['序号', '姓名', '学号', '报名时间', '签到状态', '签到时间']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        ws.row_dimensions[5].height = 30
        
        # 数据行
        for idx, apply in enumerate(applies, 1):
            row_num = idx + 5
            sign_record = sign_map.get(apply["user_id"])
            
            ws.cell(row=row_num, column=1, value=idx).alignment = cell_alignment
            ws.cell(row=row_num, column=2, value=apply.get("user_name", "")).alignment = cell_alignment
            ws.cell(row=row_num, column=3, value=apply.get("user_id", "")).alignment = cell_alignment
            
            apply_time = apply.get("apply_time", "")
            if apply_time:
                try:
                    apply_time = datetime.fromisoformat(apply_time).strftime('%Y-%m-%d %H:%M:%S')
                except:
                    pass
            ws.cell(row=row_num, column=4, value=apply_time).alignment = cell_alignment
            
            status = "已签到" if sign_record else "未签到"
            status_cell = ws.cell(row=row_num, column=5, value=status)
            status_cell.alignment = cell_alignment
            if sign_record:
                status_cell.font = Font(color='52c41a', bold=True)
            else:
                status_cell.font = Font(color='999999')
            
            sign_time = ""
            if sign_record:
                sign_time = sign_record.get("sign_time", "")
                if sign_time:
                    try:
                        sign_time = datetime.fromisoformat(sign_time).strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        pass
            ws.cell(row=row_num, column=6, value=sign_time).alignment = cell_alignment
            
            ws.row_dimensions[row_num].height = 22
        
        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 22
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        filename = f"{event.get('title', event_id)}_签到名单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        # 处理文件名中的特殊字符
        from urllib.parse import quote
        encoded_filename = quote(filename.encode('utf-8'))
        
        logger.info(f"✅ 导出签到名单: {event_id}, 共 {len(applies)} 人")
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出签到名单失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))